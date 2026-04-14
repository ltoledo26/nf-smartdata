import streamlit as st
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import os
import base64
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
import fitz  # PyMuPDF
from PIL import Image
import io
import re

# ── Config ──────────────────────────────────────────────────────────────────
EXCEL_PATH = Path("notas_fiscais.xlsx")
OPENAI_MODEL = "gpt-4o"

st.set_page_config(
    page_title="NF Smart Data",
    page_icon="🧾",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
    background-color: #0f1117;
    color: #e8eaf0;
}

.stApp { background-color: #0f1117; }

.main-header {
    background: linear-gradient(135deg, #1a1f2e 0%, #12161f 100%);
    border: 1px solid #2a3040;
    border-radius: 12px;
    padding: 2rem 2.5rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.main-header::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #4f8ef7, #7c5cbf, #4f8ef7);
    background-size: 200% 100%;
    animation: shimmer 3s linear infinite;
}
@keyframes shimmer { 0%{background-position:200% 0} 100%{background-position:-200% 0} }

.main-header h1 {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.8rem;
    font-weight: 600;
    color: #ffffff;
    margin: 0 0 0.3rem 0;
    letter-spacing: -0.5px;
}
.main-header p {
    color: #6b7a99;
    font-size: 0.9rem;
    margin: 0;
    font-weight: 300;
}

.card {
    background: #1a1f2e;
    border: 1px solid #2a3040;
    border-radius: 10px;
    padding: 1.5rem;
    margin-bottom: 1.2rem;
}
.card-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    font-weight: 600;
    color: #4f8ef7;
    text-transform: uppercase;
    letter-spacing: 2px;
    margin-bottom: 1rem;
}

.extracted-field {
    display: flex;
    align-items: flex-start;
    gap: 0.75rem;
    padding: 0.6rem 0;
    border-bottom: 1px solid #1e2535;
}
.extracted-field:last-child { border-bottom: none; }
.field-label {
    font-size: 0.75rem;
    color: #6b7a99;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    min-width: 130px;
    padding-top: 2px;
}
.field-value {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.85rem;
    color: #c8d0e0;
    word-break: break-all;
}
.field-value.not-found { color: #4a5568; font-style: italic; }

.success-banner {
    background: linear-gradient(135deg, #0d2618, #0a2010);
    border: 1px solid #1a5c30;
    border-radius: 8px;
    padding: 1rem 1.25rem;
    color: #4ade80;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.85rem;
    margin-top: 1rem;
}
.error-banner {
    background: #1e0a0a;
    border: 1px solid #5c1a1a;
    border-radius: 8px;
    padding: 1rem 1.25rem;
    color: #f87171;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.85rem;
    margin-top: 1rem;
}

.stButton > button {
    background: linear-gradient(135deg, #4f8ef7, #3a6fd8);
    color: white;
    border: none;
    border-radius: 8px;
    padding: 0.6rem 2rem;
    font-family: 'IBM Plex Sans', sans-serif;
    font-weight: 600;
    font-size: 0.9rem;
    letter-spacing: 0.3px;
    transition: all 0.2s;
    width: 100%;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #6aa3ff, #4f8ef7);
    box-shadow: 0 4px 20px rgba(79,142,247,0.3);
}

div[data-testid="stFileUploader"] {
    background: #1a1f2e;
    border: 2px dashed #2a3040;
    border-radius: 10px;
    padding: 1rem;
}
div[data-testid="stFileUploader"]:hover {
    border-color: #4f8ef7;
}

.stTextInput > div > div > input,
.stTextAreaInput > div > div > textarea {
    background: #12161f !important;
    border: 1px solid #2a3040 !important;
    border-radius: 6px !important;
    color: #e8eaf0 !important;
    font-family: 'IBM Plex Mono', monospace !important;
}
.stTextInput > div > div > input:focus,
.stTextAreaInput > div > div > textarea:focus {
    border-color: #4f8ef7 !important;
    box-shadow: 0 0 0 2px rgba(79,142,247,0.2) !important;
}

label { color: #8892a4 !important; font-size: 0.82rem !important; font-weight: 600 !important; }

.counter-badge {
    display: inline-block;
    background: #4f8ef7;
    color: white;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    font-weight: 600;
    padding: 0.2rem 0.6rem;
    border-radius: 20px;
    margin-left: 0.5rem;
}

hr { border-color: #2a3040; }
</style>
""", unsafe_allow_html=True)


# ── Excel helpers ─────────────────────────────────────────────────────────────
def get_or_create_workbook():
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notas Fiscais"
        headers = ["#", "Data de Importação", "Valor Total (R$)", "Data de Emissão",
                   "Fornecedor", "CNPJ", "Nº de Autorização", "Classe de Valor", "Natureza"]
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", start_color="1a3a6b")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="2a3040")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        widths = [5, 18, 16, 16, 35, 20, 22, 22, 22]
        for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        wb.save(EXCEL_PATH)
    return wb, wb.active


def append_row(data: dict):
    wb, ws = get_or_create_workbook()
    next_row = ws.max_row + 1
    seq = next_row - 1
    values = [
        seq,
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        data.get("valor_total", ""),
        data.get("data_emissao", ""),
        data.get("fornecedor", ""),
        data.get("cnpj", ""),
        data.get("num_autorizacao", ""),
        data.get("classe_valor", ""),
        data.get("natureza", ""),
    ]
    thin = Side(style="thin", color="2a3040")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_fill = PatternFill("solid", start_color="0f1825") if seq % 2 == 0 else PatternFill("solid", start_color="12161f")
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.font = Font(name="Arial", size=9, color="c8d0e0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = row_fill
    ws.row_dimensions[next_row].height = 20
    wb.save(EXCEL_PATH)
    return seq


def count_rows():
    if not EXCEL_PATH.exists():
        return 0
    wb, ws = get_or_create_workbook()
    return max(0, ws.max_row - 1)


# ── NF extraction helpers ─────────────────────────────────────────────────────
def parse_xml_nf(xml_bytes: bytes) -> dict:
    """Extract fields from NF-e XML."""
    try:
        root = ET.fromstring(xml_bytes)
        ns = {"nfe": "http://www.portalfiscal.inf.br/nfe"}
        def find(path):
            el = root.find(path, ns)
            return el.text.strip() if el is not None and el.text else ""
        # Try multiple paths for robustness
        valor = (find(".//nfe:vNF") or find(".//nfe:vTotTrib") or
                 find(".//nfe:vProd") or "")
        raw_date = (find(".//nfe:dhEmi") or find(".//nfe:dEmi") or "")
        if "T" in raw_date:
            raw_date = raw_date.split("T")[0]
        try:
            dt = datetime.strptime(raw_date, "%Y-%m-%d")
            date_str = dt.strftime("%d/%m/%Y")
        except:
            date_str = raw_date
        fornecedor = (find(".//nfe:emit/nfe:xNome") or find(".//nfe:emit/nfe:xFant") or "")
        cnpj = find(".//nfe:emit/nfe:CNPJ") or ""
        if cnpj:
            cnpj = re.sub(r"(\d{2})(\d{3})(\d{3})(\d{4})(\d{2})", r"\1.\2.\3/\4-\5", cnpj)
        autorizacao = (find(".//nfe:infProt/nfe:nProt") or find(".//nfe:chNFe") or "")
        if valor:
            try:
                valor = f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except:
                pass
        return {
            "valor_total": valor,
            "data_emissao": date_str,
            "fornecedor": fornecedor,
            "cnpj": cnpj,
            "num_autorizacao": autorizacao,
        }
    except Exception as e:
        return {}


def pdf_to_base64_images(pdf_bytes: bytes) -> list[str]:
    """Convert PDF pages to base64 PNG images."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    for page in doc:
        mat = fitz.Matrix(2, 2)  # 2x zoom for better OCR
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        images.append(base64.standard_b64encode(img_bytes).decode())
    return images


SYSTEM_PROMPT = """Você é um especialista em extração de dados de Notas Fiscais brasileiras.
Extraia EXATAMENTE os campos abaixo do documento fornecido e retorne SOMENTE JSON válido, sem markdown.
Campos obrigatórios:
- valor_total: valor total da NF em formato brasileiro (ex: "1.234,56"), apenas números e pontuação
- data_emissao: data no formato dd/mm/aaaa
- fornecedor: razão social ou nome do emitente
- cnpj: CNPJ do emitente no formato XX.XXX.XXX/XXXX-XX
- num_autorizacao: número de autorização, protocolo ou chave de acesso da NF

Se um campo não for encontrado, use string vazia "".
Retorne APENAS o JSON, sem explicações."""


def get_openai_client():
    api_key = os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        st.error("❌ Chave OPENAI_API_KEY não configurada. Adicione nas configurações do Streamlit Cloud.")
        st.stop()
    return OpenAI(api_key=api_key)


def extract_with_openai(messages_content: list) -> dict:
    """Call OpenAI API to extract NF fields."""
    client = get_openai_client()
    response = client.chat.completions.create(
        model=OPENAI_MODEL,
        max_tokens=1000,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": messages_content},
        ],
    )
    raw = response.choices[0].message.content.strip()
    raw = re.sub(r"```json|```", "", raw).strip()
    return json.loads(raw)


def extract_from_image(image_bytes: bytes, media_type: str) -> dict:
    b64 = base64.standard_b64encode(image_bytes).decode()
    content = [
        {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{b64}"}},
        {"type": "text", "text": "Extraia os dados desta nota fiscal."},
    ]
    return extract_with_openai(content)


def extract_from_pdf(pdf_bytes: bytes) -> dict:
    images_b64 = pdf_to_base64_images(pdf_bytes)
    content = []
    for img_b64 in images_b64[:3]:  # max 3 pages
        content.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}"}})
    content.append({"type": "text", "text": "Extraia os dados desta nota fiscal."})
    return extract_with_openai(content)


# ── UI ────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>🧾 NF Smart Data</h1>
  <p>Extração automática de Notas Fiscais → Planilha de importação</p>
</div>
""", unsafe_allow_html=True)

total = count_rows()
st.markdown(f"""
<div style="text-align:right; margin-top:-1rem; margin-bottom:1.5rem;">
  <span style="color:#6b7a99; font-size:0.8rem;">Notas registradas</span>
  <span class="counter-badge">{total}</span>
</div>
""", unsafe_allow_html=True)

# ── Step 1: Upload ────────────────────────────────────────────────────────────
st.markdown('<div class="card"><div class="card-title">① Enviar Nota Fiscal</div>', unsafe_allow_html=True)
uploaded = st.file_uploader(
    "Arraste ou clique para selecionar",
    type=["pdf", "xml", "png", "jpg", "jpeg", "webp"],
    help="Formatos aceitos: PDF, XML (NF-e), PNG, JPG"
)
st.markdown('</div>', unsafe_allow_html=True)

if uploaded:
    ext = uploaded.name.split(".")[-1].lower()
    file_bytes = uploaded.read()
    extracted = {}
    error_msg = None

    with st.spinner("🔍 Analisando nota fiscal..."):
        try:
            if ext == "xml":
                extracted = parse_xml_nf(file_bytes)
                if not any(extracted.values()):
                    # fallback to OpenAI
                    extracted = extract_with_openai([
                        {"type": "text", "text": f"Dados XML da NF:\n{file_bytes.decode('utf-8', errors='replace')}\n\nExtraia os campos."}
                    ])
            elif ext == "pdf":
                extracted = extract_from_pdf(file_bytes)
            elif ext in ("png", "jpg", "jpeg", "webp"):
                mt_map = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg", "webp": "image/webp"}
                extracted = extract_from_image(file_bytes, mt_map[ext])
        except Exception as e:
            error_msg = str(e)

    if error_msg:
        st.markdown(f'<div class="error-banner">❌ Erro na extração: {error_msg}</div>', unsafe_allow_html=True)

    # ── Step 2: Review extracted data ────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">② Dados Extraídos — Revise e Ajuste</div>', unsafe_allow_html=True)

    field_labels = {
        "valor_total":    "Valor Total (R$)",
        "data_emissao":   "Data de Emissão",
        "fornecedor":     "Fornecedor",
        "cnpj":           "CNPJ",
        "num_autorizacao":"Nº de Autorização",
    }

    col1, col2 = st.columns(2)
    inputs = {}

    fields_left  = ["valor_total", "data_emissao", "cnpj"]
    fields_right = ["fornecedor", "num_autorizacao"]

    with col1:
        for f in fields_left:
            inputs[f] = st.text_input(
                field_labels[f],
                value=extracted.get(f, ""),
                key=f"inp_{f}",
                placeholder="Não detectado — preencha manualmente"
            )
    with col2:
        for f in fields_right:
            inputs[f] = st.text_input(
                field_labels[f],
                value=extracted.get(f, ""),
                key=f"inp_{f}",
                placeholder="Não detectado — preencha manualmente"
            )

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Step 3: Classification ────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">③ Classificação Smart Data</div>', unsafe_allow_html=True)

    col3, col4 = st.columns(2)
    with col3:
        inputs["classe_valor"] = st.text_input(
            "Classe de Valor",
            key="inp_classe",
            placeholder="Ex: Despesa operacional"
        )
    with col4:
        inputs["natureza"] = st.text_input(
            "Natureza",
            key="inp_natureza",
            placeholder="Ex: Serviços"
        )
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Step 4: Save ─────────────────────────────────────────────────────────
    if st.button("💾  Salvar na Planilha", use_container_width=True):
        if not inputs.get("fornecedor") and not inputs.get("valor_total"):
            st.warning("⚠️ Preencha ao menos o Fornecedor ou o Valor Total antes de salvar.")
        else:
            seq = append_row(inputs)
            st.markdown(f"""
<div class="success-banner">
  ✅ NF #{seq} salva com sucesso em <strong>notas_fiscais.xlsx</strong><br>
  <span style="color:#86efac; font-size:0.78rem;">Fornecedor: {inputs.get('fornecedor','—')} | Valor: R$ {inputs.get('valor_total','—')}</span>
</div>""", unsafe_allow_html=True)
            st.balloons()

# ── Download button ───────────────────────────────────────────────────────────
st.markdown("<hr>", unsafe_allow_html=True)
if EXCEL_PATH.exists():
    with open(EXCEL_PATH, "rb") as f:
        excel_bytes = f.read()
    col_dl, col_info = st.columns([2, 3])
    with col_dl:
        st.download_button(
            label="⬇️  Baixar Planilha Excel",
            data=excel_bytes,
            file_name="notas_fiscais.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_info:
        st.markdown(f"""
<div style="padding:0.6rem 0; color:#6b7a99; font-size:0.8rem; line-height:1.6;">
  📁 <code style="color:#4f8ef7">notas_fiscais.xlsx</code><br>
  {count_rows()} nota(s) registrada(s)
</div>""", unsafe_allow_html=True)
else:
    st.markdown('<div style="color:#4a5568; font-size:0.8rem; text-align:center; padding:1rem;">Nenhuma nota salva ainda. Envie a primeira NF acima.</div>', unsafe_allow_html=True)
